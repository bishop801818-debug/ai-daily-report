"""
碳酸锂数据库 - 从 xlsx 生成 embedded data JS
用法: python _gen_carbonate_embedded.py
"""
import openpyxl
import json
import datetime
import re

XLSX_PATH = r"C:/Users/1/Downloads/🏔️ 【龙蟠时代】碳酸锂产业链数据库.xlsx"
OUTPUT_PATH = r"D:/trae/AI Daily report/carbonate_embedded_data.js"

def fmt_date(v):
    """将 datetime 或字符串转为 'YYYY-MM-DD 00:00:00' 格式"""
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d 00:00:00")
    if isinstance(v, str):
        # 如果已经是 YYYY-MM-DD 格式
        if len(v) == 10 and v[4] == "-" and v[7] == "-":
            return v + " 00:00:00"
        # 尝试解析
        try:
            dt = datetime.datetime.fromisoformat(v.replace("/", "-"))
            return dt.strftime("%Y-%m-%d 00:00:00")
        except:
            return v
    if v is None:
        return None
    return str(v)

def clean_val(v):
    """清理单元格值，保留原类型"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, str):
        s = v.strip()
        if s == "":
            return None
        return s
    return v

def xlsx_rows_to_records(ws, field_map, computed_fields=None):
    """
    将 worksheet 行转换为记录列表
    field_map: dict，key = 目标字段名, value = 列索引（0-based，None表示跳过）
    computed_fields: list of (field_name, formula) where formula(row) computes value
    """
    records = []
    for row_idx in range(2, ws.max_row + 1):
        row = {}
        for field, col_idx in field_map.items():
            if col_idx is None:
                continue
            cell_val = ws.cell(row_idx, col_idx + 1).value
            row[field] = clean_val(cell_val)
        # 跳过全部为空的行
        if all(v is None for v in row.values()):
            continue
        # 计算字段（千克→吨, 元→万元等）
        if computed_fields:
            for field, formula in computed_fields:
                try:
                    row[field] = formula(row)
                except Exception:
                    row[field] = None
        records.append(row)
    return records

def mk_val(v):
    """安全转换为数值"""
    if v is None: return None
    try: return float(v)
    except: return None

def sheet_to_table(sheet_idx, ws):
    """根据 sheet 索引和列数，确定表名、字段映射和计算字段"""
    n_rows = ws.max_row - 1  # 减去表头
    n_cols = ws.max_column

    # xlsx 列索引（0-based）
    C0, C1, C2, C3, C4, C5, C6, C7, C8 = 0, 1, 2, 3, 4, 5, 6, 7, 8

    if sheet_idx == 0 and n_cols == 6:
        return ("碳酸锂-行业总产量", "碳酸锂-行业总产量", {
            "日期": C0, "月份": C1, "产量（吨）": C2, "产能（吨）": C3, "产能利用率": C4, "_record_id": C5
        }, None)
    elif sheet_idx == 1 and n_cols == 9:
        return ("碳酸锂-分企业产量", "碳酸锂-分企业产量", {
            "日期": C0, "月份": C1, "企业名称": C2, "产量（吨）": C3, "环比": C4, "同比": C5,
            "产能（吨）": C6, "产能利用率": C7, "_record_id": C8
        }, None)
    elif sheet_idx == 2 and n_cols == 7:
        return ("碳酸锂-出口量", "碳酸锂-出口量", {
            "日期": C0, "月份": C1, "出口量（吨）": C2, "出口金额（元）": C3,
            "出口均价（万元/吨）": C4, "环比": C5, "同比": C6
        }, None)
    elif sheet_idx == 3 and n_cols == 8:
        # xlsx: 0=日期,1=月份,2=贸易伙伴,3=万千克(=吨),4=万元(=吨),5=元,6=万元/吨,7=元/吨
        # embedded: 出口量(万千克)=C3, 出口量(吨)=C3/1000, 出口价格(万元)=C5/10000, 出口价格(元)=C5
        return ("碳酸锂-出口贸易伙伴", "碳酸锂-出口贸易伙伴", {
            "日期": C0, "月份": C1, "贸易伙伴名称": C2,
            "出口量（万千克）": C3, "出口金额（元）": C5,
            "出口均价（万元/吨）": C6
        }, [
            ("出口量（千克）",   lambda r: mk_val(r.get("出口量（万千克）")) * 1000),
            ("出口量（吨）",     lambda r: mk_val(r.get("出口量（万千克）")) / 1000),
            ("出口价格（万元）", lambda r: mk_val(r.get("出口金额（元）")) / 10000),
            ("出口价格（元）",   lambda r: mk_val(r.get("出口金额（元）"))),
        ])
    elif sheet_idx == 4 and n_cols == 7:
        return ("碳酸锂-进口量", "碳酸锂-进口量", {
            "日期": C0, "月份": C1, "进口量（吨）": C2, "进口金额（元）": C3,
            "进口均价（万元/吨）": C4, "环比": C5, "同比": C6
        }, None)
    elif sheet_idx == 5 and n_cols == 8:
        # xlsx: 3=万千克(=吨), 4=万元(=吨), 5=元, 6=万元/吨
        return ("碳酸锂-进口贸易伙伴", "碳酸锂-进口贸易伙伴", {
            "日期": C0, "月份": C1, "贸易伙伴名称": C2,
            "进口量（万千克）": C3, "进口金额（元）": C5,
            "进口均价（万元/吨）": C6
        }, [
            ("进口量（千克）",   lambda r: mk_val(r.get("进口量（万千克）")) * 1000),
            ("进口量（吨）",     lambda r: mk_val(r.get("进口量（万千克）")) / 1000),
            ("进口价格（万元）", lambda r: mk_val(r.get("进口金额（元）")) / 10000),
            ("进口价格（元）",   lambda r: mk_val(r.get("进口金额（元）"))),
        ])
    elif sheet_idx == 6 and n_cols == 5:
        return ("碳酸锂-价格", "碳酸锂-价格", {
            "日期": C0, "规格": C1, "最低价（万元/吨）": C2, "最高价（万元/吨）": C3, "均价（万元/吨）": C4
        }, None)
    elif sheet_idx == 7 and n_cols == 5:
        return ("氢氧化锂-行业总产量", "氢氧化锂-行业总产量", {
            "日期": C0, "月份": C1, "产量（吨）": C2, "产能（吨）": C3, "产能利用率": C4
        }, None)
    elif sheet_idx == 8 and n_cols == 8:
        return ("氢氧化锂-分企业产量", "氢氧化锂-分企业产量", {
            "日期": C0, "月份": C1, "企业名称": C2, "产量（吨）": C3, "环比": C4, "同比": C5,
            "产能（吨）": C6, "产能利用率": C7
        }, None)
    elif sheet_idx == 9 and n_cols == 7:
        return ("氢氧化锂-出口量", "氢氧化锂-出口量", {
            "日期": C0, "月份": C1, "出口量（吨）": C2, "出口金额（元）": C3,
            "出口均价（万元/吨）": C4, "环比": C5, "同比": C6
        }, None)
    elif sheet_idx == 10 and n_cols == 8:
        return ("氢氧化锂-出口贸易伙伴", "氢氧化锂-出口贸易伙伴", {
            "日期": C0, "月份": C1, "贸易伙伴名称": C2,
            "出口量（万千克）": C3, "出口金额（元）": C5,
            "出口均价（万元/吨）": C6
        }, [
            ("出口量（千克）",   lambda r: mk_val(r.get("出口量（万千克）")) * 1000),
            ("出口量（吨）",     lambda r: mk_val(r.get("出口量（万千克）")) / 1000),
            ("出口价格（万元）", lambda r: mk_val(r.get("出口金额（元）")) / 10000),
            ("出口价格（元）",   lambda r: mk_val(r.get("出口金额（元）"))),
        ])
    elif sheet_idx == 11 and n_cols == 7:
        return ("氢氧化锂-进口量", "氢氧化锂-进口量", {
            "日期": C0, "月份": C1, "进口量（吨）": C2, "进口金额（元）": C3,
            "进口均价（万元/吨）": C4, "环比": C5, "同比": C6
        }, None)
    elif sheet_idx == 12 and n_cols == 8:
        return ("氢氧化锂-进口贸易伙伴", "氢氧化锂-进口贸易伙伴", {
            "日期": C0, "月份": C1, "贸易伙伴名称": C2,
            "进口量（万千克）": C3, "进口金额（元）": C5,
            "进口均价（万元/吨）": C6
        }, [
            ("进口量（千克）",   lambda r: mk_val(r.get("进口量（万千克）")) * 1000),
            ("进口量（吨）",     lambda r: mk_val(r.get("进口量（万千克）")) / 1000),
            ("进口价格（万元）", lambda r: mk_val(r.get("进口金额（元）")) / 10000),
            ("进口价格（元）",   lambda r: mk_val(r.get("进口金额（元）"))),
        ])
    elif sheet_idx == 13 and n_cols == 5:
        return ("氢氧化锂-价格", "氢氧化锂-价格", {
            "日期": C0, "规格": C1, "最低价（万元/吨）": C2, "最高价（万元/吨）": C3, "均价（万元/吨）": C4
        }, None)
    elif sheet_idx == 14 and n_cols == 5:
        return ("锂云母-价格", "锂云母-价格", {
            "日期": C0, "规格": C1, "最低价（万元/吨）": C2, "最高价（万元/吨）": C3, "均价（万元/吨）": C4
        }, None)
    elif sheet_idx == 15 and n_cols == 7:
        return ("锂辉石-进口量", "锂辉石-进口量", {
            "日期": C0, "月份": C1, "进口量（吨）": C2, "进口金额（元）": C3,
            "进口均价（元/吨）": C4, "环比": C5, "同比": C6
        }, None)
    elif sheet_idx == 16 and n_cols == 8:
        # xlsx: 3=万千克(=吨), 4=万元(=吨), 5=元, 6=元/吨
        return ("锂辉石-进口贸易伙伴", "锂辉石-进口贸易伙伴", {
            "日期": C0, "月份": C1, "贸易伙伴名称": C2,
            "进口量（万千克）": C3, "进口金额（元）": C5,
            "进口均价（元/吨）": C6
        }, [
            ("进口量（千克）",   lambda r: mk_val(r.get("进口量（万千克）")) * 1000),
            ("进口量（吨）",     lambda r: mk_val(r.get("进口量（万千克）")) / 1000),
            ("进口价格（万元）", lambda r: mk_val(r.get("进口金额（元）")) / 10000),
            ("进口价格（元）",   lambda r: mk_val(r.get("进口金额（元）"))),
        ])
    elif sheet_idx == 17 and n_cols == 7:
        return ("锂辉石-出口量", "锂辉石-出口量", {
            "日期": C0, "月份": C1, "出口量（吨）": C2, "出口金额（元）": C3,
            "出口均价（元/吨）": C4, "环比": C5, "同比": C6
        }, None)
    elif sheet_idx == 18 and n_cols == 8:
        return ("锂辉石-出口贸易伙伴", "锂辉石-出口贸易伙伴", {
            "日期": C0, "月份": C1, "贸易伙伴名称": C2,
            "出口量（万千克）": C3, "出口金额（元）": C5,
            "出口均价（万元/吨）": C6
        }, [
            ("出口量（千克）",   lambda r: mk_val(r.get("出口量（万千克）")) * 1000),
            ("出口量（吨）",     lambda r: mk_val(r.get("出口量（万千克）")) / 1000),
            ("出口价格（万元）", lambda r: mk_val(r.get("出口金额（元）")) / 10000),
            ("出口价格（元）",   lambda r: mk_val(r.get("出口金额（元）"))),
        ])
    elif sheet_idx == 19 and n_cols == 6:
        return ("锂辉石精矿-价格", "锂辉石精矿-价格", {
            "日期": C0, "规格": C1, "产地": C2, "最低价（美元/吨）": C3, "最高价（美元/吨）": C4, "均价（美元/吨）": C5
        }, None)
    else:
        return (None, None, None, None)

def sort_descending(records):
    """按日期/月份降序排列"""
    def key(r):
        d = r.get("日期") or r.get("月份") or ""
        if isinstance(d, datetime.datetime):
            return d
        if isinstance(d, str):
            # 尝试完整日期 "YYYY-MM-DD" 或 "YYYY-MM-DD HH:MM:SS"
            s = d[:10]
            try:
                return datetime.datetime.strptime(s, "%Y-%m-%d")
            except:
                pass
            # 回退到月份 "YYYY-MM"
            try:
                return datetime.datetime.strptime(d[:7], "%Y-%m")
            except:
                pass
        return datetime.datetime.min
    return sorted(records, key=key, reverse=True)

def run():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    print(f"Reading xlsx, {len(wb.sheetnames)} sheets")

    tables = []
    update_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for i, sn in enumerate(wb.sheetnames):
        ws = wb[sn]
        table_name, sheet_name, field_map, computed_fields = sheet_to_table(i, ws)

        if table_name is None:
            print(f"  [!] Sheet {i} not recognized: rows={ws.max_row}, cols={ws.max_column}, skipping")
            continue

        records = xlsx_rows_to_records(ws, field_map, computed_fields)

        # 格式化日期字段
        for r in records:
            if "日期" in r and r["日期"]:
                r["日期"] = fmt_date(r["日期"])
            if "月份" in r and r["月份"]:
                if isinstance(r["月份"], datetime.datetime):
                    r["月份"] = r["月份"].strftime("%Y-%m")
                elif isinstance(r["月份"], str) and len(r["月份"]) > 7:
                    r["月份"] = r["月份"][:7]

        # 降序排列
        records = sort_descending(records)

        # 格式化数字值
        for r in records:
            for k, v in r.items():
                if isinstance(v, float) and v != v:  # NaN
                    r[k] = None
                elif isinstance(v, float) and v == int(v) and abs(v) < 1e15:
                    r[k] = int(v)

        tables.append({
            "table_name": table_name,
            "sheet_name": sheet_name,
            "data": records,
            "row_count": len(records)
        })
        print(f"  [{i}] {table_name}: {len(records)} 条记录")

    # 写入 JS 文件
    js_content = f'''const EMBEDDED_DATA = {{
  "update_time": "{update_time}",
  "source": "🏔️ 【龙蟠时代】碳酸锂产业链数据库.xlsx",
  "tables": {json.dumps(tables, ensure_ascii=False, indent=2)}
}};
window.EMBEDDED_DATA = EMBEDDED_DATA;
'''

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        f.write(js_content)

    print(f"\n[OK] Written: {OUTPUT_PATH}")
    print(f"   Update time: {update_time}")
    total = sum(t["row_count"] for t in tables)
    print(f"   Total records: {total}")

if __name__ == "__main__":
    run()

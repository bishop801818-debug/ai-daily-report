#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
import_all.py — 7个数据库统一导入：Excel → *_embedded_data.js

用法:
  python import_all.py                 # 全部7个
  python import_all.py automotive      # 只运行指定库
  python import_all.py list            # 显示支持列表
"""
import openpyxl, json, re, os, sys, shutil, datetime, subprocess, hashlib
from datetime import datetime as dt

sys.stdout.reconfigure(encoding='utf-8')
BASE = r"D:\trae\AI Daily report"
DL   = r"C:\Users\1\Downloads"

# ── 通用工具 ─────────────────────────────────────────────────────────────────

def fmt_date(v):
    if isinstance(v, dt):
        return v.strftime("%Y-%m-%d 00:00:00")
    if isinstance(v, str):
        if len(v) == 10 and v[4] == "-" and v[7] == "-":
            return v + " 00:00:00"
        # YYYY/M/D 或 YYYY/MM/DD 格式 → 补零后转标准格式
        m_ymd = re.match(r'^(\d{4})/(\d{1,2})/(\d{1,2})$', v.strip())
        if m_ymd:
            return dt(int(m_ymd.group(1)), int(m_ymd.group(2)), int(m_ymd.group(3))).strftime("%Y-%m-%d 00:00:00")
        # M/D 或 MM/DD 格式（仅月日，无年，用当前年）
        m_md = re.match(r'^(\d{1,2})/(\d{1,2})$', v.strip())
        if m_md:
            return dt(dt.now().year, int(m_md.group(1)), int(m_md.group(2))).strftime("%Y-%m-%d 00:00:00")
        return v
    return str(v) if v is not None else None

def clean_val(v):
    if v is None: return None
    if isinstance(v, dt): return fmt_date(v)
    if isinstance(v, (int, float)):
        return None if v != v else (int(v) if v == int(v) and abs(v) < 1e15 else v)
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return v

def sort_records(records, date_key="日期"):
    def key(r):
        for k in [date_key, "当前日期", "时间", "月份", "数据年月"]:
            v = r.get(k, "") or ""
            s = str(v)[:10]
            try: return dt.strptime(s, "%Y-%m-%d")
            except: pass
            m = re.match(r"(\d{4})-(\d{2})", s)
            if m: return dt(int(m.group(1)), int(m.group(2)), 1)
        return dt.min
    return sorted(records, key=key, reverse=True)

def clean_header(h):
    """清理被 Excel 日期序列号污染的列名，如 '日期28429030' → '日期'"""
    if h is None: return None
    s = str(h).strip()
    if not s: return None
    # 日期/时间/月份类列名，尾部附加了 7-8 位数字（Excel 日期序列号）
    # 保留原始前缀，丢弃后面的数字污染部分
    m = re.match(r'^([\u4e00-\u9fff]+)\d{7,}$', s)
    if m:
        return m.group(1)
    return s

def read_sheet(ws):
    headers = [clean_header(c.value) if c.value else f"_c{i}"
               for i, c in enumerate(ws[1])]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row): continue
        obj = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                obj[headers[i]] = clean_val(val)
        if obj: records.append(obj)
    for r in records:
        for k in ["日期", "当前日期", "时间", "数据年月"]:
            if k in r and r[k]: r[k] = fmt_date(r[k])
        for k in ["月份"]:
            if k in r and r[k]:
                v = r[k]
                if isinstance(v, dt):
                    r[k] = v.strftime("%Y-%m")
                elif str(v).isdigit() and len(str(v)) >= 6:
                    # 纯数字格式如 202606 或 20260601 → 取前6位
                    r[k] = str(v)[:6]
                elif re.match(r'^\d{4}-\d{2}$', str(v)):
                    # YYYY-MM 格式
                    r[k] = str(v)
                # 中文格式如 "2026年1-6月" 或 "2026年6月" → 保持原样

    return sort_records(records)

def read_density_sheet(ws):
    """
    三行表头宽格式（LFP现货市场价-分压实密度）：
    - 第0行：密度规格名，如 磷酸铁锂（16）、磷酸铁锂（21）...（列名）
    - 第1行：类型，如 动力型 / 储能型
    - 第2行：子规格，如 2.4-2.5g/mAh、最低价、最高价、均价（仅用于分组定位）
    - 第3行起：实际数据

    实际列结构（每组5列 = 日期 + 3个价格 + 空列）：
      日期在 col0, col5, col10, col15, col20, col25, col30, col35 位置
      每组内: col+1=最低价, col+2=最高价, col+3=均价
    输出：标准扁平记录，每条含 日期/规格/类型/最低价/最高价/均价
    """
    rows = list(ws.iter_rows(values_only=True))
    header_row = rows[0]   # 密度规格名（列名）
    type_row   = rows[1]   # 动力型/储能型
    spec_row   = rows[2]   # 2.4-2.5g/mAh / 最低价 / 最高价 / 均价
    raw_data   = rows[3:]  # 数据从第4行开始

    # 找每组的起始列（密度规格列，即 spec_row 中有 'g/mAh' 的位置）
    # 每组结构: [密度规格(来自header), 日期(来自row), 最低价, 最高价, 均价, 空]
    # 对应列: i, i+1, i+2, i+3, i+4 (每组5列)
    group_starts = []
    for i in range(len(spec_row)):
        v = clean_val(spec_row[i])
        if v and isinstance(v, str) and 'g/mAh' in v:
            group_starts.append(i)

    records = []
    for row in raw_data:
        if all(v is None for v in row):
            continue
        date_raw = clean_val(row[0])
        if not date_raw:
            continue
        date_str = fmt_date(date_raw)

        for gi, start in enumerate(group_starts):
            # 该组的密度规格名取自第0行列名
            density_name = clean_val(header_row[start])
            # 该组的类型取自第1行
            dtype = clean_val(type_row[start]) or "动力型"
            # 价格列: start+1=最低价, start+2=最高价, start+3=均价
            if start + 3 >= len(row):
                continue
            lo = clean_val(row[start + 1])
            hi = clean_val(row[start + 2])
            avg = clean_val(row[start + 3])

            records.append({
                "日期": date_str,
                "规格": str(density_name) if density_name else "-",
                "类型": str(dtype),
                "最低价（万元/吨）": clean_val(lo),
                "最高价（万元/吨）": clean_val(hi),
                "均价（万元/吨）": clean_val(avg),
            })

    return sort_records(records)

def xlsx_to_js(excel_path, sheets_cfg, output_js, source_name, key=None):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    tables = []
    for cfg in sheets_cfg:
        idx   = cfg["idx"]
        tname = cfg["table_name"]
        extra = cfg.get("computed", None)

        ws = wb.worksheets[idx]
        records = cfg.get("read_fn", read_sheet)(ws)
        if extra:
            for r in records:
                for fname, formula in extra:
                    try: r[fname] = formula(r)
                    except: r[fname] = None

        tables.append({
            "table_name": tname,
            "sheet_name": wb.sheetnames[idx],
            "data": records,
            "row_count": len(records)
        })
        print(f"    [{idx:2d}] {tname}: {len(records)} rows")

    wb.close()
    update_time = dt.now().strftime("%Y-%m-%d %H:%M:%S")
    js = f'''const EMBEDDED_DATA = {{
  "update_time": "{update_time}",
  "source": "{source_name}",
  "tables": {json.dumps(tables, ensure_ascii=False, indent=2)}
}};
window.EMBEDDED_DATA = EMBEDDED_DATA;
'''
    with open(output_js, "w", encoding="utf-8") as f:
        f.write(js)
    print(f"  → {os.path.basename(output_js)}  ({os.path.getsize(output_js):,} bytes)")

    # 成功后记录指纹
    md5, rows = _get_excel_fingerprint(excel_path, sheets_cfg)
    fp = _load_fingerprint()
    fp[key] = {"md5": md5, "rows": rows, "ts": dt.now().strftime("%Y-%m-%d %H:%M:%S")}
    _save_fingerprint(fp)
    print(f"  ✓ 指纹已记录: MD5={md5[:8]}")

    return True

# ── 7个数据库配置 ────────────────────────────────────────────────────────────

def sh(idx, tname, computed=None):
    return {"idx": idx, "table_name": tname, "computed": computed}

CONFIGS = {

    "automotive": {
        "label": "🚗 汽车",
        "excel": os.path.join(DL, "🚗全球汽车市场数据库 (1).xlsx"),
        "output": os.path.join(BASE, "automotive_embedded_data.js"),
        "source": "🚗全球汽车市场数据库.xlsx",
        "use_existing": True,
        "existing_script": os.path.join(BASE, "_regen_automotive.py"),
        "sheets": None,
    },

    "carbonate": {
        "label": "🏔️ 碳酸锂",
        "excel": os.path.join(DL, "🏔️ 【龙蟠时代】碳酸锂产业链数据库.xlsx"),
        "output": os.path.join(BASE, "carbonate_embedded_data.js"),
        "source": "🏔️ 【龙蟠时代】碳酸锂产业链数据库.xlsx",
        "use_existing": True,
        "existing_script": os.path.join(BASE, "_gen_carbonate_embedded.py"),
        "sheets": None,
    },

    "electrolyte": {
        "label": "🧪 电解液",
        "excel": os.path.join(DL, "电解液行业数据库.xlsx"),
        "output": os.path.join(BASE, "electrolyte_embedded_data.js"),
        "source": "电解液行业数据库.xlsx",
        "sheets": [
            sh(0,"电解液-行业整体产量"), sh(1,"电解液-分企业产量横向"),
            sh(2,"电解液-top15排名"), sh(3,"电解液年累-top15"),
            sh(4,"电解液价格-磷酸铁锂动力型"), sh(5,"电解液价格-磷酸铁锂储能型"),
            sh(6,"电解液价格-三元动力型"), sh(7,"电解液价格-圆柱2600mAh"),
            sh(8,"电解液价格-圆柱2200mAh"), sh(9,"高压电解液价格-4.4V以上"),
            sh(10,"高压电解液价格-4.4V"), sh(11,"高压电解液价格-4.35V"),
            sh(12,"电解液价格-分企业横向"), sh(13,"六氟磷酸锂-行业总产量"),
            sh(14,"六氟-分企业产量"), sh(15,"六氟-top15排名"),
            sh(16,"六氟年累-top15"), sh(17,"六氟磷酸锂价格-主流市场"),
            sh(18,"六氟磷酸锂价格-出口"), sh(19,"LiFSI价格-固态"),
            sh(20,"LiFSI价格-液态"), sh(21,"六氟磷酸锂出口-总量"),
            sh(22,"六氟出口-分国别"), sh(23,"添加剂VC-产量"),
            sh(24,"添加剂FEC-产量"), sh(25,"添加剂VC-价格"),
            sh(26,"添加剂PS-价格"), sh(27,"添加剂FEC-价格"),
        ],
    },

    "lfp": {
        "label": "🔋 磷酸铁锂",
        "excel": os.path.join(DL, "🔋 【常州锂源】磷酸铁锂产业链数据库.xlsx"),
        "output": os.path.join(BASE, "lfp_embedded_data.js"),
        "source": "🔋 【常州锂源】磷酸铁锂产业链数据库.xlsx",
        "sheets": [
            sh(0,"LFP-行业整体产量"), sh(1,"LFP-分企业产量"),
            sh(2,"LFP竞对销量"), sh(3,"LFP-出口量"),
            sh(4,"LFP-出口贸易伙伴"), sh(5,"FP-行业整体产量"),
            sh(6,"FP-分企业产量"), sh(7,"FP竞对销量"),
            # Sheet 8 = 宽格式分压实密度（自定义三行表头读取逻辑）
            {"idx": 8, "table_name": "LFP现货市场价-分压实密度",
             "read_fn": read_density_sheet},
            sh(9,"LFP现货市场价"), sh(10,"FP现货市场价"),
            sh(11,"磷酸盐价格"), sh(12,"非磷原料价格"),
        ],
    },

    "lib_battery": {
        "label": "🔋 锂电池",
        "excel": os.path.join(DL, "🔋 2.0 锂电池行业数据库.xlsx"),
        "output": os.path.join(BASE, "lib_battery_embedded_data.js"),
        "source": "🔋 2.0 锂电池行业数据库.xlsx",
        "sheets": [
            sh(0,"锂电池行业产量-分规格"), sh(1,"锂电池行业产量产能"),
            sh(2,"锂电池分企业产量-分规格"), sh(3,"锂电池分企业产量产能"),
            sh(4,"电池排产预测"),
        ],
    },

    "ternary": {
        "label": "🔋 三元前驱体",
        "excel": os.path.join(DL, "🔋 【三金锂电】三元前驱体产业链数据库.xlsx"),
        "output": os.path.join(BASE, "ternary_embedded_data.js"),
        "source": "🔋 【三金锂电】三元前驱体产业链数据库.xlsx",
        "sheets": [
            sh(0,"NCM-行业整体产量"), sh(1,"NCM-分型号产量"),
            sh(2,"NCM-分企业产量"), sh(3,"NCM-出口总量"),
            sh(4,"NCM-出口目的地"), sh(5,"NCM-进口总量"),
            sh(6,"NCM-进口目的地"), sh(7,"NCM-现货市场价"),
            sh(8,"NCMpre-行业整体产量"), sh(9,"NCMpre-分型号产量"),
            sh(10,"NCMpre-分企业产量"), sh(11,"NCMpre-出口总量"),
            sh(12,"NCMpre-出口目的地"), sh(13,"NCMpre-进口总量"),
            sh(14,"NCMpre-进口目的地"), sh(15,"NCMpre-现货市场价"),
            sh(16,"关键原料-现货市场价"), sh(17,"四氧化三钴-行业整体产量"),
            sh(18,"四氧化三钴-现货市场价"),
        ],
    },

    "recycling": {
        "label": "♻️ 回收",
        "excel": os.path.join(DL, "锂电池回收行业数据库.xlsx"),
        "output": os.path.join(BASE, "recycling_embedded_data.js"),
        "source": "锂电池回收行业数据库.xlsx",
        "sheets": [
            sh(0,"黑粉处理量-总计"), sh(1,"黑粉处理量-三元+钴酸锂"),
            sh(2,"黑粉处理量-铁锂"), sh(3,"黑粉整体处理量-分企业"),
            sh(4,"黑粉三元+钴酸锂处理量-分企业"), sh(5,"黑粉磷酸铁锂处理量-分企业"),
            sh(6,"锂电池价格-大三元聚合物电池包"), sh(7,"锂电池价格-大三元铝壳电池包"),
            sh(8,"锂电池价格-钴酸锂铝壳电池包"), sh(9,"锂电池价格-加工费电池包"),
            sh(10,"锂电池价格-铁锂钢壳电池包"), sh(11,"锂电池价格-铁锂铝壳电池包"),
            sh(12,"极片价格-负极片"), sh(13,"极片价格-钴酸锂正极片"),
            sh(14,"极片价格-加工费"), sh(15,"极片价格-三元523正极片"),
            sh(16,"极片价格-铁锂正极片"), sh(17,"辅料价格-铝粉"),
            sh(18,"辅料价格-铜粉"), sh(19,"黑粉价格-钴酸锂电池粉"),
            sh(20,"黑粉价格-钴酸锂极片粉"), sh(21,"黑粉价格-三元523电池粉"),
            sh(22,"黑粉价格-三元523极片粉"), sh(23,"黑粉价格-石墨粉"),
            sh(24,"黑粉价格-铁锂电池粉加工费"), sh(25,"黑粉价格-铁锂极片粉"),
            sh(26,"黑粉价格-铁锂极片粉加工费"),
        ],
    },
}

# ── Excel 变更指纹追踪 ────────────────────────────────────────────────────────
# 记录每个 Excel 的 MD5 + 关键 sheet 行数，运行前对比，有异常变化时告警
FINGERPRINT_FILE = os.path.join(BASE, "_excel_fingerprint.json")

def _load_fingerprint():
    if os.path.exists(FINGERPRINT_FILE):
        try:
            with open(FINGERPRINT_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {}
    return {}

def _save_fingerprint(fp):
    with open(FINGERPRINT_FILE, 'w', encoding='utf-8') as f:
        json.dump(fp, f, ensure_ascii=False, indent=2)

def _get_excel_fingerprint(excel_path, sheets_cfg):
    """计算 Excel 的 MD5 + 每个 sheet 的行数（作为指纹）"""
    md5 = hashlib.md5(open(excel_path, 'rb').read()).hexdigest()
    row_counts = []
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
        for cfg in (sheets_cfg or []):
            idx = cfg["idx"]
            if idx < len(wb.worksheets):
                ws = wb.worksheets[idx]
                count = sum(1 for _ in ws.iter_rows())
                row_counts.append(count)
            else:
                row_counts.append(-1)
        wb.close()
    except Exception:
        pass
    return md5, row_counts

def _check_and_warn(key, excel_path, sheets_cfg):
    """运行前检测 Excel 是否有异常变化，返回 (是否变化, 旧指纹)"""
    fp = _load_fingerprint()
    prev = fp.get(key, {})
    old_md5 = prev.get("md5")
    old_rows = prev.get("rows", [])
    new_md5, new_rows = _get_excel_fingerprint(excel_path, sheets_cfg)

    changed = (old_md5 != new_md5)

    # 对比每个 sheet 行数（只对比前几个，避免空 sheet 误报）
    rows_changed = False
    if old_md5 == new_md5 and old_rows and new_rows:
        # MD5 没变则不需要检查行数
        pass
    elif old_md5 != new_md5 and old_rows and new_rows:
        # MD5 变了，检查主 sheet（第0个）行数是否大幅下降
        if old_rows and new_rows and len(old_rows) > 0 and len(new_rows) > 0:
            delta = old_rows[0] - new_rows[0]
            if delta > 5:  # 主 sheet 行数突然少了5行以上 → 可能被覆盖
                rows_changed = True

    if changed or rows_changed:
        old_ts = prev.get("ts", "?")
        print(f"\n  ⚠️  【Excel 变更检测】")
        if changed:
            print(f"     文件 MD5 变化: {old_md5[:8] if old_md5 else '无'} → {new_md5[:8]}")
        if rows_changed:
            print(f"     主 sheet 行数下降: {old_rows[0]} → {new_rows[0]} （少了 {delta} 行）")
            print(f"     上次记录时间: {old_ts}")
            print(f"     ⚠️  Excel 可能被覆盖为旧版本，数据可能不完整！")
        else:
            print(f"     上次记录时间: {old_ts}")

    return changed or rows_changed, fp

# ── Excel 自动存档 ────────────────────────────────────────────────────────────
SOURCES_DIR = os.path.join(BASE, "_sources")

def _archive_excel(excel_path, key, label):
    """成功运行后将 Excel 复制到 _sources/ 目录并 git add"""
    if not os.path.exists(excel_path):
        return
    name = os.path.basename(excel_path)
    dst = os.path.join(SOURCES_DIR, name)
    try:
        shutil.copy2(excel_path, dst)
        # git add
        subprocess.run(["git", "add", f"_sources/{name}"],
                       cwd=BASE, capture_output=True,
                       encoding="utf-8", errors="replace")
        print(f"  📦 Excel 已存档: _sources/{name}")
    except Exception as e:
        print(f"  [WARN] Excel 存档失败: {e}")

# ── 运行 ─────────────────────────────────────────────────────────────────────

def run_db(key):
    cfg = CONFIGS[key]
    print(f"\n{'═'*56}")
    print(f"  {cfg['label']}  ({key})")
    print(f"{'═'*56}")

    if not os.path.exists(cfg["excel"]):
        print(f"  [SKIP] Excel 不存在:\n    {cfg['excel']}")
        return False

    # 运行前检测 Excel 变更
    sheets_cfg = cfg.get("sheets") if not (cfg.get("use_existing") and cfg.get("existing_script")) else None
    _check_and_warn(key, cfg["excel"], sheets_cfg)

    # 已有专用脚本 → 直接调用
    if cfg.get("use_existing") and cfg.get("existing_script"):
        script = cfg["existing_script"]
        if os.path.exists(script):
            print(f"  → 调用: {os.path.basename(script)}")
            r = subprocess.run([sys.executable, script],
                             capture_output=True, text=True,
                             encoding="utf-8", errors="replace")
            print(r.stdout)
            if r.returncode == 0:
                # 成功后记录指纹（专用脚本也用同一份 Excel）
                md5, _ = _get_excel_fingerprint(cfg["excel"], None)
                fp = _load_fingerprint()
                fp[key] = {"md5": md5, "rows": [], "ts": dt.now().strftime("%Y-%m-%d %H:%M:%S")}
                _save_fingerprint(fp)
                print(f"  ✓ 指纹已记录: MD5={md5[:8]}")
                _archive_excel(cfg["excel"], key, cfg["label"])
                print(f"  [OK]")
                return True
            else:
                print(f"  [WARN] 脚本返回: {r.returncode}")
                if r.stderr: print(r.stderr[:300])
                return False

    # 通用读取路径
    print(f"  Excel: {os.path.basename(cfg['excel'])}")
    print(f"  Sheets:")
    try:
        xlsx_to_js(cfg["excel"], cfg["sheets"], cfg["output"], cfg["source"], key=key)
        _archive_excel(cfg["excel"], key, cfg["label"])
        return True
    except Exception as e:
        print(f"  [ERROR] {e}")
        import traceback; traceback.print_exc()
        return False

def sync_embedded(src):
    dst = os.path.join(BASE, "embedded", os.path.basename(src))
    if os.path.exists(src):
        shutil.copy2(src, dst)
        print(f"    [SYNC] → embedded/{os.path.basename(src)}")

def main():
    args = sys.argv[1:]
    if "list" in args or (not args):
        print("\n支持数据库:")
        for k, v in CONFIGS.items():
            ok = "✅" if os.path.exists(v["excel"]) else "❌"
            print(f"  {k:16s} {v['label']}  {ok}  {os.path.basename(v['excel'])}")
        print("\n用法: python import_all.py [db_key ...]")
        return

    targets = args if args else list(CONFIGS.keys())
    print(f"{'═'*56}")
    print(f"  import_all.py — 7数据库统一导入  {dt.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"{'═'*56}")

    results = {}
    for key in targets:
        if key not in CONFIGS:
            print(f"[ERROR] 未知: {key}")
            continue
        results[key] = run_db(key)

    ok = sum(results.values())
    print(f"\n{'═'*56}")
    print(f"  完成: {ok}/{len(results)} 成功")
    print(f"{'═'*56}")

    # 同步 embedded/
    print("\n同步 embedded/:")
    for key, ok in results.items():
        if ok:
            sync_embedded(CONFIGS[key]["output"])

    # bust cache
    for cand in [r"C:\Users\1\bust_cache.py",
                 os.path.join(BASE, "..", "bust_cache.py"),
                 "bust_cache.py"]:
        if os.path.exists(cand):
            print("\n更新缓存戳:")
            r = subprocess.run([sys.executable, cand],
                             capture_output=True, text=True,
                             encoding="utf-8", errors="replace")
            print(r.stdout)
            break

    print("\n✅ 完成 — Ctrl+Shift+R 验证")

if __name__ == "__main__":
    main()
